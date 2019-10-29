from openpyxl import load_workbook
import glob

# TO-DO-List:
# TODO: Fix variable-names to normal standards
# All English speaking people should be able to read it ;)

# This Excel file contains all information on all students...
csreport = 'report.e88414.01-10-2019.xlsx'

# Import the Excel to a Workbook...
wb = load_workbook(filename = csreport)
sheet_ranges = wb['Overview']

def readrow(rownr):
    mylist = []
    for row in sheet_ranges.iter_cols(min_col=rownr, max_col=rownr, min_row=2, max_row=None, values_only=False):
        for cell in row:
            mylist.append(cell.value)
    return mylist

# Lets create students...

allstudents = readrow(0)

for students in allstudents:
    student = {'email': students[]}

